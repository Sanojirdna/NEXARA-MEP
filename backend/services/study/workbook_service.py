from __future__ import annotations

from collections import defaultdict
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from services.technikraum.technikraum_vdi_service import TECHNIKRAUM_VDI

from .study_utils import (
    SELECTABLE_COLUMNS,
    SELECTED_FILL,
    SHEET_NAME,
    STUDY_EXPORT_PATH,
    STUDY_TEMPLATE_PATH,
    TABLE_END_COL,
    TABLE_END_ROW,
    TABLE_START_COL,
    TABLE_START_ROW,
    WEIGHT_COLUMN,
    StudyUtilityMixin,
)


class StudyWorkbookService(StudyUtilityMixin):
    """Build the Studie frontend table and Excel workbook exports."""

    def build_table(self, ws: Any, selected_coords: set[str], weights: dict[str, int]) -> list[list[dict[str, Any]]]:
        """Return table rows for the Studie page."""
        return self._build_table(ws, selected_coords, weights)

    def build_selection_summary(self, ws: Any, selected_coords: set[str], weights: dict[str, int]) -> dict[str, Any]:
        """Return the selected Studie rows grouped by criterion."""
        return self._build_selection_summary(ws, selected_coords, weights)

    def ensure_export_workbook(self, state: dict[str, Any]) -> Path:
        """Create or update the Studie export workbook."""
        return self._ensure_export_workbook(state)

    def export_workbook(self, runtime: Any, state: dict[str, Any]) -> Path:
        """Export the Studie workbook with runtime sheets."""
        workbook_path = self._ensure_export_workbook(state)
        self._write_runtime_sheets(workbook_path, runtime, state)
        return workbook_path

    def _build_table(self, ws: Any, selected_coords: set[str], weights: dict[str, int]) -> list[list[dict[str, Any]]]:
        merge_info, covered_cells = self._get_merge_info(ws)
        table_rows: list[list[dict[str, Any]]] = []

        for row_idx in range(TABLE_START_ROW, TABLE_END_ROW + 1):
            row_cells: list[dict[str, Any]] = []
            for col_idx in range(TABLE_START_COL, TABLE_END_COL + 1):
                if (row_idx, col_idx) in covered_cells:
                    continue

                cell = ws.cell(row_idx, col_idx)
                span = merge_info.get((row_idx, col_idx), {"rowspan": 1, "colspan": 1})
                coord = cell.coordinate
                criterion = self._criterion_from_row(row_idx, ws)
                row_cells.append(
                    {
                        "coord": coord,
                        "row": row_idx,
                        "col": col_idx,
                        "value": "" if cell.value is None else cell.value,
                        "rowspan": span["rowspan"],
                        "colspan": span["colspan"],
                        "css_class": self._row_css_class(row_idx, ws),
                        "selectable": (
                            col_idx in SELECTABLE_COLUMNS
                            and row_idx not in {32, 66}
                            and criterion is not None
                            and bool(str(cell.value or "").strip())
                        ),
                        "selected": coord in selected_coords,
                        "weight_editable": col_idx == WEIGHT_COLUMN and criterion is not None,
                        "weight_selected": f"F{row_idx}" in weights,
                        "weight_value": weights.get(f"F{row_idx}", ""),
                        "criterion": criterion,
                    }
                )
            table_rows.append(row_cells)

        return table_rows

    def _build_selection_summary(self, ws: Any, selected_coords: set[str], weights: dict[str, int]) -> dict[str, Any]:
        selections_by_criterion: dict[int, list[dict[str, Any]]] = defaultdict(list)
        total_score = 0

        for coord in sorted(selected_coords, key=self._sort_coord_key):
            row_idx = self._row_from_coord(coord)
            criterion = self._criterion_from_row(row_idx, ws)
            if criterion is None:
                continue

            code = str(ws[f"B{row_idx}"].value or "").replace("\u200b", "").strip()
            label = str(ws[coord].value or "").strip()
            weight = weights.get(f"F{row_idx}")
            if isinstance(weight, int):
                total_score += weight

            selections_by_criterion[criterion].append(
                {
                    "coord": coord,
                    "row": row_idx,
                    "code": code,
                    "label": label,
                    "weight": weight,
                }
            )

        score_band = self._score_band(total_score)
        return {
            "total_score": total_score,
            "score_band": score_band,
            "score_band_label": self._score_band_label(score_band),
            "required_share_label": self._required_share_label(score_band),
            "selections_by_criterion": dict(sorted(selections_by_criterion.items())),
        }

    def _ensure_export_workbook(self, state: dict[str, Any]) -> Path:
        wb = load_workbook(STUDY_TEMPLATE_PATH)
        ws = wb[SHEET_NAME]

        for row_idx in range(7, 66):
            ws[f"F{row_idx}"] = None

        # reset fills in selectable area by copying template fills from a fresh workbook
        template_wb = load_workbook(STUDY_TEMPLATE_PATH)
        template_ws = template_wb[SHEET_NAME]
        for row_idx in range(7, 66):
            for col_idx in SELECTABLE_COLUMNS:
                coord = f"{self._column_letter(col_idx)}{row_idx}"
                ws[coord]._style = copy(template_ws[coord]._style)

        selected_coords = set(state.get("selected_coords", []))
        weights = self._normalized_weights(state.get("weights", {}))
        total = 0
        for coord in selected_coords:
            ws[coord].fill = copy(SELECTED_FILL)

        for coord, value in weights.items():
            ws[coord] = value
            total += int(value)

        ws["F66"] = total
        STUDY_EXPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
        wb.save(STUDY_EXPORT_PATH)
        return STUDY_EXPORT_PATH

    def _write_runtime_sheets(self, workbook_path: Path, runtime: Any, state: dict[str, Any]) -> None:
        wb = load_workbook(workbook_path)

        self._replace_sheet(wb, "Project_Info")
        ws_info = wb["Project_Info"]
        info_rows = [
            ("IFC file", getattr(runtime, "current_ifc_name", None)),
            ("Demand Excel", getattr(runtime, "current_excel_name", None)),
            ("Config file", getattr(runtime, "current_config_name", None)),
            ("Config source", getattr(runtime, "current_config_source", None)),
            ("Has bundle", runtime.bundle is not None),
            ("Loaded from bundle", bool((runtime.bundle or {}).get("loaded_from_bundle"))),
            ("Workers", getattr(runtime.current_config, "default_workers", None)),
            ("Voxel size", getattr(runtime.current_config, "voxel_size", None)),
            ("Candidate shaft limit", getattr(runtime.current_config, "candidate_shaft_limit", None)),
            ("Studie selected coords", ", ".join(state.get("selected_coords", []))),
        ]
        for row_index, (label, value) in enumerate(info_rows, start=1):
            ws_info[f"A{row_index}"] = label
            ws_info[f"B{row_index}"] = value

        self._replace_sheet(wb, "Planner_Config")
        ws_config = wb["Planner_Config"]
        ws_config.append(["section", "key", "value"])
        config_payload = {
            "voxel_grid.voxel_size": getattr(runtime.current_config, "voxel_size", None),
            "voxel_grid.voxel_margin": getattr(runtime.current_config.penalty_config, "voxel_margin", None),
            "voxel_grid.route_clearance_margin": getattr(runtime.current_config.penalty_config, "route_clearance_margin", None),
            "runtime.default_workers": getattr(runtime.current_config, "default_workers", None),
            "runtime.candidate_shaft_limit": getattr(runtime.current_config, "candidate_shaft_limit", None),
            "penalties.wall_cross_penalty": getattr(runtime.current_config.penalty_config, "wall_cross_penalty", None),
            "penalties.slab_cross_penalty": getattr(runtime.current_config.penalty_config, "slab_cross_penalty", None),
            "penalties.blocked_penalty": getattr(runtime.current_config.penalty_config, "blocked_penalty", None),
            "penalties.wall_distance_clip": getattr(runtime.current_config.penalty_config, "wall_distance_clip", None),
            "penalties.corridor_distance_clip": getattr(runtime.current_config.penalty_config, "corridor_distance_clip", None),
        }
        for key, value in config_payload.items():
            section, short_key = key.split(".", 1)
            ws_config.append([section, short_key, value])

        for keyword in getattr(runtime.current_config.keyword_config, "corridor_keywords", []):
            ws_config.append(["keywords.corridor", "value", keyword])
        for keyword in getattr(runtime.current_config.keyword_config, "shaft_keywords", []):
            ws_config.append(["keywords.shaft", "value", keyword])
        for keyword in getattr(runtime.current_config.keyword_config, "no_route_space_keywords", []):
            ws_config.append(["keywords.no_route", "value", keyword])

        for service, allowed in getattr(runtime.current_config, "shaft_allow_map", {}).items():
            ws_config.append(["shaft_allow_map", str(service), ", ".join(allowed)])

        for strategy_name, strategy in getattr(runtime.current_config, "strategies", {}).items():
            strategy_dict = strategy.to_dict()
            for key, value in strategy_dict.items():
                ws_config.append([f"strategy.{strategy_name}", key, value])

        self._replace_sheet(wb, "Room_Demands")
        ws_demands = wb["Room_Demands"]
        ws_demands.append([
            "demand_id",
            "room_guid",
            "room_name",
            "service",
            "media_type",
            "hvac_system_type",
            "kind",
            "value",
            "unit",
        ])
        for demand in (runtime.bundle or {}).get("demands", []):
            row = demand.to_dict()
            ws_demands.append([
                row.get("demand_id"),
                row.get("room_guid"),
                row.get("room_name"),
                row.get("service"),
                row.get("media_type"),
                row.get("hvac_system_type"),
                row.get("kind"),
                row.get("value"),
                row.get("unit"),
            ])

        self._replace_sheet(wb, "Selected_Variants")
        ws_selected = wb["Selected_Variants"]
        ws_selected.append([
            "demand_id",
            "room_name",
            "service",
            "shaft_guid",
            "shaft_name",
            "strategy",
            "success",
            "score",
            "length_m",
            "bend_count",
            "wall_crossings",
            "slab_crossings",
        ])
        for route in ((runtime.current_system.to_dict() if runtime.current_system else {}).get("routes", [])):
            metrics = route.get("metrics", {}) or {}
            ws_selected.append([
                route.get("demand_id"),
                route.get("room_name"),
                route.get("service"),
                route.get("shaft_guid"),
                route.get("shaft_name"),
                route.get("strategy"),
                route.get("success"),
                route.get("score"),
                metrics.get("length_m"),
                metrics.get("bend_count"),
                metrics.get("wall_crossings"),
                metrics.get("slab_crossings"),
            ])

        TECHNIKRAUM_VDI.write_workbook_sheet(wb, runtime, study_state=state)
        wb.save(workbook_path)

    def _normalized_weights(self, raw_weights: dict[str, Any]) -> dict[str, int]:
        normalized: dict[str, int] = {}
        for key, value in (raw_weights or {}).items():
            coord = str(key or "").strip().upper()
            if not coord.startswith("F"):
                continue
            if value in (None, ""):
                continue
            try:
                parsed = int(value)
            except Exception:
                continue
            if parsed in (1, 2, 3):
                normalized[coord] = parsed
        return normalized

    def _score_band(self, total_score: int) -> str | None:
        if total_score <= 0:
            return None
        if total_score <= 5:
            return "low"
        if total_score <= 10:
            return "medium"
        return "high"

    def _score_band_label(self, score_band: str | None) -> str | None:
        labels = {
            "low": "gering",
            "medium": "mittel",
            "high": "hoch",
        }
        return labels.get(score_band or "")

    def _required_share_label(self, score_band: str | None) -> str | None:
        labels = {
            "low": "kleiner Schachtflächenanteil",
            "medium": "mittlerer Schachtflächenanteil",
            "high": "hoher Schachtflächenanteil",
        }
        return labels.get(score_band or "")

    def _replace_sheet(self, wb: Any, sheet_name: str) -> None:
        if sheet_name in wb.sheetnames:
            ws_old = wb[sheet_name]
            wb.remove(ws_old)
        wb.create_sheet(sheet_name)
