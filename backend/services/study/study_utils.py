from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


BACKEND_DIR = Path(__file__).resolve().parents[2]
PROJECT_ROOT = BACKEND_DIR.parent
UPLOADS_DIR = PROJECT_ROOT / "uploads"
STUDY_TEMPLATE_PATH = BACKEND_DIR / "study_templates" / "diagramm_1_1.xlsx"
STUDY_STATE_PATH = UPLOADS_DIR / "studie_state.json"
STUDY_EXPORT_PATH = UPLOADS_DIR / "studie_export.xlsx"

SHEET_NAME = "Diagramm 1.1"
SELECTABLE_COLUMNS = {3, 4, 5}  # C, D, E
WEIGHT_COLUMN = 6  # F
TABLE_START_ROW = 1
TABLE_END_ROW = 69
TABLE_START_COL = 2  # B
TABLE_END_COL = 7  # G
SELECTED_FILL = PatternFill(fill_type="solid", fgColor="D7F4D2")


class StudyUtilityMixin:
    """Shared helper methods for the Studie services."""

    def _get_merge_info(self, ws: Any) -> tuple[dict[tuple[int, int], dict[str, int]], set[tuple[int, int]]]:
        top_left_cells: dict[tuple[int, int], dict[str, int]] = {}
        covered_cells: set[tuple[int, int]] = set()

        for merged in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged.bounds
            top_left_cells[(min_row, min_col)] = {
                "rowspan": max_row - min_row + 1,
                "colspan": max_col - min_col + 1,
            }
            for row_idx in range(min_row, max_row + 1):
                for col_idx in range(min_col, max_col + 1):
                    if (row_idx, col_idx) != (min_row, min_col):
                        covered_cells.add((row_idx, col_idx))

        return top_left_cells, covered_cells

    def _criterion_from_row(self, row_idx: int, ws: Any | None = None) -> int | None:
        workbook = None
        if ws is None:
            workbook = load_workbook(STUDY_TEMPLATE_PATH, read_only=True)
            ws = workbook[SHEET_NAME]
        code = str(ws[f"B{row_idx}"].value or "").replace("\u200b", "").strip()
        match = re.match(r"^(\d+)\.\d+", code)
        if workbook is not None:
            workbook.close()
        return int(match.group(1)) if match else None

    def _criterion_from_coord(self, coord: str) -> int | None:
        return self._criterion_from_row(self._row_from_coord(coord))

    def _row_css_class(self, row_idx: int, ws: Any) -> str:
        if row_idx <= 5:
            return "header"
        if row_idx == 66:
            return "sum"
        if row_idx >= 67:
            return "note"
        b_value = str(ws[f"B{row_idx}"].value or "").replace("\u200b", "").strip()
        if re.match(r"^\d+\s", b_value):
            return "section"
        return ""

    def _row_from_coord(self, coord: str) -> int:
        match = re.search(r"(\d+)$", coord)
        return int(match.group(1)) if match else 0

    def _col_from_coord(self, coord: str) -> int:
        letters = re.match(r"([A-Z]+)", coord.upper())
        if not letters:
            return 0
        result = 0
        for char in letters.group(1):
            result = result * 26 + (ord(char) - 64)
        return result

    def _sort_coord_key(self, coord: str) -> tuple[int, int]:
        return (self._row_from_coord(coord), self._col_from_coord(coord))

    def _column_letter(self, index: int) -> str:
        letters = ""
        value = index
        while value > 0:
            value, remainder = divmod(value - 1, 26)
            letters = chr(65 + remainder) + letters
        return letters

    def _union_area(self, rects: list[tuple[float, float, float, float]]) -> float:
        if not rects:
            return 0.0

        x_coords = sorted({x0 for x0, _, x1, _ in rects} | {x1 for _, _, x1, _ in rects})
        total = 0.0
        for idx in range(len(x_coords) - 1):
            x_left = x_coords[idx]
            x_right = x_coords[idx + 1]
            if x_right <= x_left:
                continue

            y_intervals: list[tuple[float, float]] = []
            for rx0, ry0, rx1, ry1 in rects:
                if rx0 <= x_left and rx1 >= x_right:
                    y_intervals.append((ry0, ry1))
            if not y_intervals:
                continue

            y_intervals.sort()
            merged: list[tuple[float, float]] = []
            current_start, current_end = y_intervals[0]
            for start, end in y_intervals[1:]:
                if start <= current_end:
                    current_end = max(current_end, end)
                else:
                    merged.append((current_start, current_end))
                    current_start, current_end = start, end
            merged.append((current_start, current_end))

            height = sum(max(0.0, end - start) for start, end in merged)
            total += (x_right - x_left) * height
        return total
